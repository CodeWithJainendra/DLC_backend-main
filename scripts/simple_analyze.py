#!/usr/bin/env python3

"""
Simple script to analyze Excel files and determine how to import them
"""

import os
from openpyxl import load_workbook

def analyze_excel_file(file_path):
    """Analyze an Excel file and print its structure"""
    print(f"\n{'='*80}")
    print(f"Analyzing file: {file_path}")
    print(f"{'='*80}")
    
    if not os.path.exists(file_path):
        print(f"❌ File not found: {file_path}")
        return
    
    try:
        # Load workbook
        wb = load_workbook(file_path, read_only=True)
        print(f"✅ File loaded successfully")
        print(f"📋 Sheets: {wb.sheetnames}")
        
        # Analyze each sheet
        for sheet_name in wb.sheetnames[:1]:  # Only analyze first sheet
            print(f"\n📄 Sheet: {sheet_name}")
            ws = wb[sheet_name]
            
            # Get dimensions
            print(f"   Dimensions: {ws.max_row} rows × {ws.max_column} columns")
            
            # Show column headers (first row)
            if ws.max_row > 0:
                headers = []
                for cell in next(ws.iter_rows(max_col=min(20, ws.max_column), values_only=True)):
                    headers.append(str(cell) if cell is not None else "")
                print(f"   Headers: {headers}")
                
    except Exception as e:
        print(f"❌ Error analyzing file: {e}")

def main():
    # List of files to analyze
    files_to_analyze = [
        "BOB Pensioners data 1.xlsx",
        "BOB Pensioners data 2.xlsx",
        "Dashborad_DLC_Data_.xlsx",
        "Data from UBI 1.xlsx",
        "Data from UBI 2.xlsx",
        "Data from UBI 3.xlsx"
    ]
    
    base_path = "/data1/jainendra/DLC_backend-main"
    
    print("🔍 ANALYZING EXCEL FILES FOR IMPORT")
    print("="*80)
    
    for file_name in files_to_analyze:
        file_path = os.path.join(base_path, file_name)
        analyze_excel_file(file_path)

if __name__ == "__main__":
    main()