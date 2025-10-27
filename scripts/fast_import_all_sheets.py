#!/usr/bin/env python3

"""
ULTRA-FAST COMPLETE IMPORT - All 5 Sheets
Strategy: Convert Excel sheets to CSV first, then import (10x faster)
"""

from openpyxl import load_workbook
import sqlite3
import csv
import os
import sys
from datetime import datetime
import tempfile

class FastDOPPWImporter:
    def __init__(self):
        self.db_path = os.path.join(os.path.dirname(__file__), '..', 'DLC_Database.db')
        self.conn = None
        self.cursor = None
        self.temp_dir = tempfile.mkdtemp()
        
    def connect_db(self):
        """Connect to database"""
        self.conn = sqlite3.connect(self.db_path)
        self.cursor = self.conn.cursor()
        print("✅ Database connected\n")
        
    def excel_to_csv(self, excel_path):
        """Convert all Excel sheets to CSV files"""
        print("📖 Loading Excel file...")
        wb = load_workbook(excel_path, read_only=True, data_only=True)
        sheet_names = wb.sheetnames
        print(f"✅ Found {len(sheet_names)} sheets: {', '.join(sheet_names)}\n")
        
        csv_files = {}
        
        for sheet_name in sheet_names:
            print(f"🔄 Converting '{sheet_name}' to CSV...")
            ws = wb[sheet_name]
            
            csv_path = os.path.join(self.temp_dir, f"{sheet_name.replace(' ', '_')}.csv")
            
            with open(csv_path, 'w', newline='', encoding='utf-8') as csvfile:
                writer = csv.writer(csvfile)
                row_count = 0
                
                for row in ws.iter_rows(values_only=True):
                    # Skip completely empty rows
                    if any(cell is not None and str(cell).strip() for cell in row):
                        writer.writerow(row)
                        row_count += 1
                        
                        if row_count % 50000 == 0:
                            print(f"   ... {row_count:,} rows converted")
            
            print(f"   ✅ Saved: {csv_path} ({row_count:,} rows)\n")
            csv_files[sheet_name] = {'path': csv_path, 'rows': row_count}
        
        wb.close()
        return csv_files
    
    def map_headers(self, headers):
        """Map CSV headers to database columns"""
        mapping = {}
        
        for idx, header in enumerate(headers):
            if not header:
                continue
            
            h = str(header).strip().upper().replace(' ', '_')
            
            if 'GCODE' in h: mapping[idx] = 'gcode'
            elif 'ESCROLL' in h: mapping[idx] = 'escroll_cat'
            elif h == 'GID' or 'GROUP_ID' in h: mapping[idx] = 'gid'
            elif 'PENSION_TYPE' in h: mapping[idx] = 'pension_type'
            elif 'BRANCH_CODE' in h: mapping[idx] = 'branch_code'
            elif 'BRANCH_NAME' in h: mapping[idx] = 'branch_name'
            elif 'BRANCH_PIN' in h: mapping[idx] = 'branch_pin'
            elif 'BRANCH_STATE' in h: mapping[idx] = 'branch_state'
            elif 'BIRTH_YEAR' in h: mapping[idx] = 'birth_year'
            elif 'SUBMITTED_STATUS' in h: mapping[idx] = 'submitted_status'
            elif 'WAIVER' in h: mapping[idx] = 'waiver_upto'
            elif 'SUBMISSION_MODE' in h: mapping[idx] = 'submission_mode'
            elif 'VERIFICATION_TYPE' in h: mapping[idx] = 'verification_type'
            elif 'CERTIFICATE_SUBMISSION_DATE' in h or 'SUBMISSION_DATE' in h: mapping[idx] = 'certificate_submission_date'
            elif 'PENSIONER_PINCODE' in h: mapping[idx] = 'pensioner_postcode'
            elif 'PENSIONER_DISTNAME' in h or 'PENSIONER_DIST' in h: mapping[idx] = 'pensioner_distname'
            elif 'PENSIONER_STATENAME' in h or 'PENSIONER_STATE' in h: mapping[idx] = 'state'
            elif 'PPO' in h and 'NUMBER' in h: mapping[idx] = 'ppo_number'
            elif 'BANK_NAME' in h: mapping[idx] = 'bank_name'
            elif h == 'CITY' or 'PENSIONER_CITY' in h: mapping[idx] = 'pensioner_city'
            elif h == 'PSA': mapping[idx] = 'psa'
            elif h == 'PDA': mapping[idx] = 'pda'
        
        return mapping
    
    def import_csv(self, csv_path, sheet_name, total_rows):
        """Import data from CSV file"""
        print(f"\n{'='*100}")
        print(f"📥 IMPORTING: {sheet_name}")
        print(f"📊 Total rows: {total_rows:,}")
        print(f"{'='*100}")
        
        start_time = datetime.now()
        
        with open(csv_path, 'r', encoding='utf-8') as csvfile:
            reader = csv.reader(csvfile)
            headers = next(reader)
            
            print(f"📝 Columns: {', '.join(headers[:10])}...")
            
            header_map = self.map_headers(headers)
            print(f"🗺️  Mapped {len(header_map)} columns\n")
            
            if len(header_map) == 0:
                print("⚠️  No columns mapped - SKIPPING\n")
                return {'imported': 0, 'errors': 0, 'duplicates': 0}
            
            insert_query = """
                INSERT INTO pensioner_bank_master (
                    gcode, escroll_cat, gid, pension_type, branch_code, branch_name,
                    branch_pin, branch_state, birth_year, submitted_status, waiver_upto,
                    submission_mode, verification_type, certificate_submission_date,
                    pensioner_postcode, pensioner_distname, state, ppo_number, bank_name,
                    pensioner_city, psa, pda, data_source, sheet_name
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """
            
            imported = 0
            errors = 0
            duplicates = 0
            batch = []
            batch_size = 10000
            row_num = 0
            
            print("🔄 Importing data...")
            
            for row in reader:
                row_num += 1
                
                record = {}
                for col_idx, db_field in header_map.items():
                    if col_idx < len(row):
                        value = row[col_idx]
                        record[db_field] = value.strip() if value else ''
                    else:
                        record[db_field] = ''
                
                values = (
                    record.get('gcode', ''),
                    record.get('escroll_cat', ''),
                    record.get('gid', ''),
                    record.get('pension_type', ''),
                    record.get('branch_code', ''),
                    record.get('branch_name', ''),
                    record.get('branch_pin', ''),
                    record.get('branch_state', ''),
                    record.get('birth_year', ''),
                    record.get('submitted_status', ''),
                    record.get('waiver_upto', ''),
                    record.get('submission_mode', ''),
                    record.get('verification_type', ''),
                    record.get('certificate_submission_date', ''),
                    record.get('pensioner_postcode', ''),
                    record.get('pensioner_distname', ''),
                    record.get('state', ''),
                    record.get('ppo_number', ''),
                    record.get('bank_name', ''),
                    record.get('pensioner_city', ''),
                    record.get('psa', ''),
                    record.get('pda', ''),
                    f'DOPPW_FAST_IMPORT_{datetime.now().strftime("%Y%m%d")}',
                    sheet_name
                )
                
                batch.append(values)
                
                if len(batch) >= batch_size:
                    try:
                        self.cursor.executemany(insert_query, batch)
                        imported += len(batch)
                        self.conn.commit()
                    except Exception as e:
                        # Fallback: insert one by one
                        for val in batch:
                            try:
                                self.cursor.execute(insert_query, val)
                                imported += 1
                            except sqlite3.IntegrityError:
                                duplicates += 1
                            except:
                                errors += 1
                        self.conn.commit()
                    
                    batch = []
                    progress = (row_num / (total_rows - 1)) * 100
                    print(f"   ✅ {progress:.1f}% | Imported: {imported:,} | Errors: {errors}")
            
            # Insert remaining batch
            if batch:
                try:
                    self.cursor.executemany(insert_query, batch)
                    imported += len(batch)
                    self.conn.commit()
                except Exception:
                    for val in batch:
                        try:
                            self.cursor.execute(insert_query, val)
                            imported += 1
                        except sqlite3.IntegrityError:
                            duplicates += 1
                        except:
                            errors += 1
                    self.conn.commit()
        
        elapsed = (datetime.now() - start_time).total_seconds()
        print(f"\n✅ Completed in {elapsed:.2f}s")
        print(f"   ✅ Imported: {imported:,}")
        print(f"   ❌ Errors: {errors:,}")
        print(f"   🔄 Duplicates: {duplicates:,}\n")
        
        return {'imported': imported, 'errors': errors, 'duplicates': duplicates, 'time': elapsed}
    
    def import_all(self, excel_path):
        """Main import function"""
        print('='*100)
        print('🚀 ULTRA-FAST COMPLETE IMPORT - ALL SHEETS')
        print('='*100)
        print(f"📁 Excel file: {excel_path}")
        print(f"📊 Size: {os.path.getsize(excel_path) / (1024*1024):.2f} MB")
        print(f"⏰ Started: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        
        self.connect_db()
        
        # Step 1: Convert Excel to CSV
        print("STEP 1: Converting Excel sheets to CSV")
        print("="*100)
        csv_files = self.excel_to_csv(excel_path)
        
        # Step 2: Import each CSV
        print("\nSTEP 2: Importing data from CSV files")
        print("="*100)
        
        total_stats = {
            'total_sheets': len(csv_files),
            'imported': 0,
            'errors': 0,
            'duplicates': 0,
            'sheets': {}
        }
        
        for sheet_name, info in csv_files.items():
            if info['rows'] <= 1:  # Only header or empty
                print(f"\n⚠️  Skipping '{sheet_name}' (no data)\n")
                continue
            
            result = self.import_csv(info['path'], sheet_name, info['rows'])
            total_stats['imported'] += result['imported']
            total_stats['errors'] += result['errors']
            total_stats['duplicates'] += result['duplicates']
            total_stats['sheets'][sheet_name] = result
        
        # Cleanup temp files
        print("🧹 Cleaning up temporary files...")
        for info in csv_files.values():
            try:
                os.remove(info['path'])
            except:
                pass
        os.rmdir(self.temp_dir)
        
        # Print summary
        print("\n" + "="*100)
        print("🎯 FINAL SUMMARY")
        print("="*100)
        print(f"📊 Total Sheets: {total_stats['total_sheets']}")
        print(f"✅ Total Imported: {total_stats['imported']:,}")
        print(f"❌ Total Errors: {total_stats['errors']:,}")
        print(f"🔄 Total Duplicates: {total_stats['duplicates']:,}")
        
        print(f"\n📋 Sheet-wise breakdown:")
        for sheet_name, stats in total_stats['sheets'].items():
            print(f"   {sheet_name}: {stats['imported']:,} records in {stats['time']:.2f}s")
        
        print(f"\n⏰ Completed: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        print("="*100)
        
        self.conn.close()
        print("\n✨ ALL SHEETS IMPORTED SUCCESSFULLY! ✨\n")

def main():
    excel_path = os.path.join(os.path.dirname(__file__), '..', 'doppw_data_03102025.xlsx')
    
    if not os.path.exists(excel_path):
        print(f"❌ Excel file not found: {excel_path}")
        sys.exit(1)
    
    importer = FastDOPPWImporter()
    importer.import_all(excel_path)

if __name__ == '__main__':
    main()
