#!/usr/bin/env python3

"""
COMPLETE DOPPW DATA IMPORT - ALL 5 SHEETS
Uses openpyxl for efficient large Excel file handling
Imports all sheets: Export Worksheet, Sheet1, Sheet2, Sheet3, Sheet4
"""

from openpyxl import load_workbook
import sqlite3
import sys
import os
from datetime import datetime

class DOPPWImporter:
    def __init__(self):
        self.db_path = os.path.join(os.path.dirname(__file__), '..', 'DLC_Database.db')
        self.conn = None
        self.cursor = None
        self.stats = {
            'total_sheets': 0,
            'processed_sheets': 0,
            'total_records': 0,
            'imported_records': 0,
            'error_records': 0,
            'duplicate_records': 0,
            'sheet_details': {}
        }

    def connect_db(self):
        """Connect to SQLite database"""
        self.conn = sqlite3.connect(self.db_path)
        self.cursor = self.conn.cursor()
        print("✅ Database connected")

    def map_headers(self, headers):
        """Map Excel headers to database columns"""
        mapping = {}
        
        for idx, header in enumerate(headers):
            if not header:
                continue
            
            h = str(header).strip().upper().replace(' ', '_')
            
            if 'GCODE' in h:
                mapping[idx] = 'gcode'
            elif 'ESCROLL' in h:
                mapping[idx] = 'escroll_cat'
            elif h == 'GID' or 'GROUP_ID' in h:
                mapping[idx] = 'gid'
            elif 'PENSION_TYPE' in h or h == 'PENSION_TYPE':
                mapping[idx] = 'pension_type'
            elif h == 'BRANCH_CODE' or 'BRANCH_CODE' in h:
                mapping[idx] = 'branch_code'
            elif h == 'BRANCH_NAME' or 'BRANCH_NAME' in h:
                mapping[idx] = 'branch_name'
            elif h == 'BRANCH_PIN' or 'BRANCH_PIN' in h:
                mapping[idx] = 'branch_pin'
            elif h == 'BRANCH_STATE' or 'BRANCH_STATE' in h:
                mapping[idx] = 'branch_state'
            elif 'BIRTH_YEAR' in h or 'YEAR_OF_BIRTH' in h:
                mapping[idx] = 'birth_year'
            elif 'SUBMITTED_STATUS' in h or h == 'SUBMITTED_STATUS':
                mapping[idx] = 'submitted_status'
            elif 'WAIVER' in h:
                mapping[idx] = 'waiver_upto'
            elif 'SUBMISSION_MODE' in h:
                mapping[idx] = 'submission_mode'
            elif 'VERIFICATION_TYPE' in h:
                mapping[idx] = 'verification_type'
            elif 'CERTIFICATE_SUBMISSION_DATE' in h or 'SUBMISSION_DATE' in h:
                mapping[idx] = 'certificate_submission_date'
            elif 'PENSIONER_PINCODE' in h:
                mapping[idx] = 'pensioner_postcode'
            elif 'PENSIONER_DISTNAME' in h or 'PENSIONER_DIST' in h:
                mapping[idx] = 'pensioner_distname'
            elif 'PENSIONER_STATENAME' in h or 'PENSIONER_STATE' in h:
                mapping[idx] = 'state'
            elif 'PPO' in h and 'NUMBER' in h:
                mapping[idx] = 'ppo_number'
            elif 'BANK_NAME' in h:
                mapping[idx] = 'bank_name'
            elif h == 'CITY' or 'PENSIONER_CITY' in h:
                mapping[idx] = 'pensioner_city'
            elif h == 'PSA':
                mapping[idx] = 'psa'
            elif h == 'PDA':
                mapping[idx] = 'pda'
        
        return mapping

    def import_sheet(self, ws, sheet_name, sheet_number):
        """Import data from a single sheet"""
        print(f"\n{'='*100}")
        print(f"📄 SHEET {sheet_number}: \"{sheet_name}\"")
        print(f"{'='*100}")
        
        start_time = datetime.now()
        
        try:
            # Get sheet dimensions
            max_row = ws.max_row
            max_col = ws.max_column
            
            print(f"📐 Sheet dimensions: {max_row:,} rows × {max_col} columns")
            
            if max_row <= 1:
                print("⚠️  No data rows - SKIPPING")
                self.stats['sheet_details'][sheet_name] = {'status': 'empty', 'records': 0}
                return
            
            # Read headers (first row)
            headers = []
            for col in range(1, max_col + 1):
                cell_value = ws.cell(row=1, column=col).value
                headers.append(cell_value if cell_value else '')
            
            print(f"📝 Found {len(headers)} columns: {', '.join(str(h) for h in headers[:10])}...")
            
            # Create header mapping
            header_map = self.map_headers(headers)
            print(f"🗺️  Mapped {len(header_map)} columns to database fields")
            
            if len(header_map) == 0:
                print("⚠️  No columns could be mapped - SKIPPING")
                self.stats['sheet_details'][sheet_name] = {'status': 'no_mapping', 'records': 0}
                return
            
            # Show sample data from first data row
            print(f"\n📋 Sample data from first row:")
            for col_idx, db_field in list(header_map.items())[:10]:
                value = ws.cell(row=2, column=col_idx + 1).value
                if value:
                    print(f"   {headers[col_idx]}: {value}")
            
            # Prepare insert query
            insert_query = """
                INSERT INTO pensioner_bank_master (
                    gcode, escroll_cat, gid, pension_type, branch_code, branch_name,
                    branch_pin, branch_state, birth_year, submitted_status, waiver_upto,
                    submission_mode, verification_type, certificate_submission_date,
                    pensioner_postcode, pensioner_distname, state, ppo_number, bank_name,
                    pensioner_city, psa, pda, data_source, sheet_name
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """
            
            # Import data in batches
            data_rows = max_row - 1  # Exclude header
            print(f"\n📥 Importing {data_rows:,} records...")
            
            batch_size = 5000
            total_batches = (data_rows + batch_size - 1) // batch_size
            imported = 0
            errors = 0
            duplicates = 0
            
            for batch_num in range(total_batches):
                start_row = 2 + (batch_num * batch_size)
                end_row = min(start_row + batch_size, max_row + 1)
                
                batch_data = []
                for row_num in range(start_row, end_row):
                    # Extract data based on header mapping
                    record = {}
                    for col_idx, db_field in header_map.items():
                        value = ws.cell(row=row_num, column=col_idx + 1).value
                        record[db_field] = str(value).strip() if value else ''
                    
                    # Prepare values for insertion
                    values = (
                        record.get('gcode', ''),
                        record.get('escroll_cat', ''),
                        record.get('gid', ''),
                        record.get('pension_type', ''),
                        record.get('branch_code', ''),
                        record.get('branch_name', ''),
                        record.get('branch_pin', ''),
                        record.get('branch_state', ''),
                        record.get('birth_year', ''),
                        record.get('submitted_status', ''),
                        record.get('waiver_upto', ''),
                        record.get('submission_mode', ''),
                        record.get('verification_type', ''),
                        record.get('certificate_submission_date', ''),
                        record.get('pensioner_postcode', ''),
                        record.get('pensioner_distname', ''),
                        record.get('state', ''),
                        record.get('ppo_number', ''),
                        record.get('bank_name', ''),
                        record.get('pensioner_city', ''),
                        record.get('psa', ''),
                        record.get('pda', ''),
                        f'DOPPW_PYTHON_IMPORT_{datetime.now().strftime("%Y%m%d")}',
                        sheet_name
                    )
                    batch_data.append(values)
                
                # Use executemany for faster bulk insert
                try:
                    self.cursor.executemany(insert_query, batch_data)
                    imported += len(batch_data)
                    self.conn.commit()
                except Exception as e:
                    # If bulk insert fails, try one by one
                    for values in batch_data:
                        try:
                            self.cursor.execute(insert_query, values)
                            imported += 1
                        except sqlite3.IntegrityError:
                            duplicates += 1
                        except Exception:
                            errors += 1
                    self.conn.commit()
                
                progress = ((batch_num + 1) / total_batches) * 100
                if (batch_num + 1) % 10 == 0 or batch_num == 0:  # Print every 10 batches
                    print(f"   📦 Batch {batch_num + 1}/{total_batches} | {progress:.1f}% | ✅ {imported:,} | ❌ {errors}")
            
            elapsed = (datetime.now() - start_time).total_seconds()
            
            print(f"\n✅ Sheet completed in {elapsed:.2f}s")
            print(f"   ✅ Imported: {imported:,}")
            print(f"   ❌ Errors: {errors:,}")
            print(f"   🔄 Duplicates: {duplicates:,}")
            
            self.stats['sheet_details'][sheet_name] = {
                'status': 'completed',
                'records': data_rows,
                'imported': imported,
                'errors': errors,
                'duplicates': duplicates,
                'time': elapsed
            }
            self.stats['processed_sheets'] += 1
            self.stats['total_records'] += data_rows
            self.stats['imported_records'] += imported
            self.stats['error_records'] += errors
            self.stats['duplicate_records'] += duplicates
            
        except Exception as e:
            print(f"❌ Error processing sheet: {str(e)}")
            self.stats['sheet_details'][sheet_name] = {
                'status': 'error',
                'error': str(e)
            }

    def import_all_sheets(self, excel_path):
        """Import all sheets from Excel file"""
        print('='*100)
        print('🚀 COMPLETE DOPPW DATA IMPORT - ALL SHEETS (Python/openpyxl)')
        print('='*100)
        print(f"📁 File: {excel_path}")
        
        file_size = os.path.getsize(excel_path) / (1024 * 1024)
        print(f"📊 Size: {file_size:.2f} MB")
        print(f"⏰ Started: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        
        # Connect to database
        self.connect_db()
        
        # Load workbook
        print("📖 Loading Excel file (this may take a moment for large files)...")
        wb = load_workbook(excel_path, read_only=True, data_only=True)
        
        sheet_names = wb.sheetnames
        self.stats['total_sheets'] = len(sheet_names)
        
        print(f"✅ File loaded! Found {len(sheet_names)} sheets")
        print(f"📋 Sheets: {', '.join(sheet_names)}\n")
        
        # Process each sheet
        for idx, sheet_name in enumerate(sheet_names, 1):
            ws = wb[sheet_name]
            self.import_sheet(ws, sheet_name, idx)
        
        # Close workbook
        wb.close()
        
        # Print summary
        self.print_summary()
        
        # Close database
        self.conn.close()

    def print_summary(self):
        """Print final import summary"""
        print('\n\n' + '='*100)
        print('🎯 FINAL SUMMARY - ALL SHEETS IMPORT')
        print('='*100)
        
        print('\n📋 SHEET-WISE RESULTS:\n')
        for sheet, details in self.stats['sheet_details'].items():
            print(f"📄 {sheet}:")
            print(f"   Status: {details['status'].upper()}")
            if details['status'] == 'completed':
                print(f"   Records: {details['records']:,}")
                print(f"   ✅ Imported: {details['imported']:,}")
                print(f"   ❌ Errors: {details['errors']:,}")
                print(f"   🔄 Duplicates: {details['duplicates']:,}")
                print(f"   ⏱️  Time: {details['time']:.2f}s")
            elif details.get('error'):
                print(f"   Error: {details['error']}")
            print()
        
        print('='*100)
        print('📊 OVERALL TOTALS:')
        print(f"   📋 Total Sheets: {self.stats['total_sheets']}")
        print(f"   ✅ Processed: {self.stats['processed_sheets']}")
        print(f"   📊 Total Records: {self.stats['total_records']:,}")
        print(f"   ✅ Imported: {self.stats['imported_records']:,}")
        print(f"   ❌ Errors: {self.stats['error_records']:,}")
        print(f"   🔄 Duplicates: {self.stats['duplicate_records']:,}")
        
        if self.stats['total_records'] > 0:
            success_rate = (self.stats['imported_records'] / self.stats['total_records']) * 100
            print(f"   📈 Success Rate: {success_rate:.2f}%")
        
        print('\n🔍 VERIFICATION COMMANDS:')
        print('   sqlite3 DLC_Database.db "SELECT COUNT(*) FROM pensioner_bank_master WHERE data_source LIKE \'DOPPW_PYTHON_IMPORT_%\';"')
        print('   sqlite3 DLC_Database.db "SELECT sheet_name, COUNT(*) FROM pensioner_bank_master WHERE data_source LIKE \'DOPPW_PYTHON_IMPORT_%\' GROUP BY sheet_name;"')
        
        print(f"\n⏰ Completed: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        print('='*100)

def main():
    excel_path = os.path.join(os.path.dirname(__file__), '..', 'doppw_data_03102025.xlsx')
    
    if not os.path.exists(excel_path):
        print(f"❌ Excel file not found: {excel_path}")
        sys.exit(1)
    
    importer = DOPPWImporter()
    importer.import_all_sheets(excel_path)
    
    print("\n✨ IMPORT COMPLETED SUCCESSFULLY! ✨\n")

if __name__ == '__main__':
    main()
