#!/usr/bin/env python3

"""
Script to analyze Excel files and determine how to import them into the pensioner_bank_master table
"""

import pandas as pd
import os
from openpyxl import load_workbook

def analyze_excel_file(file_path):
    """Analyze an Excel file and print its structure"""
    print(f"\n{'='*80}")
    print(f"Analyzing file: {file_path}")
    print(f"{'='*80}")
    
    if not os.path.exists(file_path):
        print(f"❌ File not found: {file_path}")
        return
    
    try:
        # Load workbook
        wb = load_workbook(file_path, read_only=True)
        print(f"✅ File loaded successfully")
        print(f"📋 Sheets: {wb.sheetnames}")
        
        # Analyze each sheet
        for sheet_name in wb.sheetnames[:2]:  # Only analyze first 2 sheets to avoid too much output
            print(f"\n📄 Sheet: {sheet_name}")
            ws = wb[sheet_name]
            
            # Get dimensions
            print(f"   Dimensions: {ws.max_row} rows × {ws.max_column} columns")
            
            # Show first few rows
            print("   First 5 rows:")
            for row_num, row in enumerate(ws.iter_rows(values_only=True), 1):
                if row_num > 5:
                    break
                print(f"     Row {row_num}: {row[:10]}")  # Show first 10 columns
                
            # Show column headers
            if ws.max_row > 0:
                headers = [cell.value for cell in next(ws.iter_rows(values_only=True))]
                print(f"   Headers: {headers[:15]}")  # Show first 15 headers
                
    except Exception as e:
        print(f"❌ Error analyzing file: {e}")

def main():
    # List of files to analyze
    files_to_analyze = [
        "BOB Pensioners data 1.xlsx",
        "BOB Pensioners data 2.xlsx",
        "Dashborad_DLC_Data_.xlsx",
        "Data from UBI 1.xlsx",
        "Data from UBI 2.xlsx",
        "Data from UBI 3.xlsx"
    ]
    
    base_path = "/data1/jainendra/DLC_backend-main"
    
    print("🔍 ANALYZING EXCEL FILES FOR IMPORT")
    print("="*80)
    
    for file_name in files_to_analyze:
        file_path = os.path.join(base_path, file_name)
        analyze_excel_file(file_path)

if __name__ == "__main__":
    main()