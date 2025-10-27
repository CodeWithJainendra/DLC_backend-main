#!/usr/bin/env python3

"""
Script to import additional Excel files into the pensioner_bank_master table
"""

import sqlite3
import os
from openpyxl import load_workbook
from datetime import datetime

class AdditionalDataImporter:
    def __init__(self, db_path):
        self.db_path = db_path
        self.conn = None
        self.cursor = None
        
    def connect_db(self):
        """Connect to the database"""
        self.conn = sqlite3.connect(self.db_path)
        self.cursor = self.conn.cursor()
        print("✅ Database connected")
        
    def close_db(self):
        """Close database connection"""
        if self.conn:
            self.conn.close()
            print("✅ Database connection closed")
            
    def import_bob_data(self, file_path, source_name):
        """Import BOB Pensioners data"""
        print(f"\n🔄 Importing {source_name} from {file_path}")
        
        if not os.path.exists(file_path):
            print(f"❌ File not found: {file_path}")
            return 0
            
        try:
            wb = load_workbook(file_path, read_only=True)
            ws = wb['Sheet1']
            
            # Mapping for BOB data
            headers = [str(cell) if cell is not None else "" for cell in next(ws.iter_rows(values_only=True))]
            print(f"📋 Headers: {headers}")
            
            # Create mapping from column headers to database fields
            header_map = {}
            for idx, header in enumerate(headers):
                header_clean = header.strip().upper()
                if header_clean == 'SR NO':
                    header_map[idx] = 'sr_no'
                elif header_clean == 'PPO NUMBER':
                    header_map[idx] = 'ppo_number'
                elif header_clean == 'DOB REGULAR':
                    header_map[idx] = 'pensioner_dob'
                elif header_clean == 'PSA':
                    header_map[idx] = 'psa'
                elif 'PDA' in header_clean and 'DISBURSING' in header_clean:
                    header_map[idx] = 'pda'
                elif header_clean == 'BRANCH_NAME':
                    header_map[idx] = 'branch_name'
                elif 'BRANCH' in header_clean and 'POST' in header_clean:
                    header_map[idx] = 'branch_postcode'
                elif 'PENSIONER' in header_clean and 'CITY' in header_clean:
                    header_map[idx] = 'pensioner_city'
                elif header_clean == 'STATE':
                    header_map[idx] = 'state'
                elif 'PENSIONER' in header_clean and 'POST' in header_clean:
                    header_map[idx] = 'pensioner_postcode'
            
            print(f"🗺️  Column mapping: {header_map}")
            
            # Prepare insert statement
            insert_sql = """
                INSERT INTO pensioner_bank_master (
                    sr_no, ppo_number, pensioner_dob, psa, pda, branch_name, branch_postcode,
                    pensioner_city, state, pensioner_postcode, data_source, sheet_name
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """
            
            # Process data rows
            imported_count = 0
            error_count = 0
            
            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
                try:
                    # Create record with mapped values
                    record = {
                        'sr_no': '',
                        'ppo_number': '',
                        'pensioner_dob': '',
                        'psa': '',
                        'pda': '',
                        'branch_name': '',
                        'branch_postcode': '',
                        'pensioner_city': '',
                        'state': '',
                        'pensioner_postcode': ''
                    }
                    
                    # Map values from Excel row
                    for col_idx, field_name in header_map.items():
                        if col_idx < len(row) and row[col_idx] is not None:
                            record[field_name] = str(row[col_idx]).strip()
                    
                    # Insert into database
                    values = (
                        record['sr_no'],
                        record['ppo_number'],
                        record['pensioner_dob'],
                        record['psa'],
                        record['pda'],
                        record['branch_name'],
                        record['branch_postcode'],
                        record['pensioner_city'],
                        record['state'],
                        record['pensioner_postcode'],
                        source_name,
                        'Sheet1'
                    )
                    
                    self.cursor.execute(insert_sql, values)
                    imported_count += 1
                    
                    # Show progress
                    if imported_count % 5000 == 0:
                        print(f"   ✅ Imported {imported_count} records...")
                        self.conn.commit()
                        
                except Exception as e:
                    error_count += 1
                    if error_count < 10:  # Only show first 10 errors
                        print(f"   ❌ Error importing row {row_num}: {e}")
            
            self.conn.commit()
            print(f"✅ Completed {source_name} import: {imported_count} records imported, {error_count} errors")
            return imported_count
            
        except Exception as e:
            print(f"❌ Error importing {source_name}: {e}")
            return 0
            
    def import_dlc_data(self, file_path, source_name):
        """Import Dashboard DLC data"""
        print(f"\n🔄 Importing {source_name} from {file_path}")
        
        if not os.path.exists(file_path):
            print(f"❌ File not found: {file_path}")
            return 0
            
        try:
            wb = load_workbook(file_path, read_only=True)
            ws = wb['Export Worksheet']
            
            # Mapping for DLC data
            headers = [str(cell) if cell is not None else "" for cell in next(ws.iter_rows(values_only=True))]
            print(f"📋 Headers: {headers}")
            
            # Create mapping from column headers to database fields
            header_map = {}
            for idx, header in enumerate(headers):
                header_clean = header.strip().upper()
                if header_clean == 'GCODE':
                    header_map[idx] = 'gcode'
                elif header_clean == 'ESCROLL_CAT':
                    header_map[idx] = 'escroll_cat'
                elif header_clean == 'GID':
                    header_map[idx] = 'gid'
                elif header_clean == 'PENSION_TYPE':
                    header_map[idx] = 'pension_type'
                elif header_clean == 'BRANCH_CODE':
                    header_map[idx] = 'branch_code'
                elif header_clean == 'BRANCH_NAME':
                    header_map[idx] = 'branch_name'
                elif header_clean == 'BRANCH_PIN':
                    header_map[idx] = 'branch_postcode'
                elif header_clean == 'BRANCH_STATE':
                    header_map[idx] = 'branch_state'
                elif header_clean == 'BIRTH_YEAR':
                    header_map[idx] = 'birth_year'
                elif header_clean == 'SUBMITTED_STATUS':
                    header_map[idx] = 'submitted_status'
                elif header_clean == 'WAIVER_UPTO':
                    header_map[idx] = 'waiver_upto'
                elif header_clean == 'SUBMISSION_MODE':
                    header_map[idx] = 'submission_mode'
                elif header_clean == 'VERIFICATION_TYPE':
                    header_map[idx] = 'verification_type'
                elif header_clean == 'CERTIFICATE_SUBMISSION_DATE':
                    header_map[idx] = 'certificate_submission_date'
                elif 'PENSIONER' in header_clean and 'PINCODE' in header_clean:
                    header_map[idx] = 'pensioner_postcode'
                elif 'PENSIONER' in header_clean and 'DIST' in header_clean:
                    header_map[idx] = 'pensioner_distname'
                elif 'PENSIONER' in header_clean and 'STATE' in header_clean:
                    header_map[idx] = 'state'
            
            print(f"🗺️  Column mapping: {header_map}")
            
            # Prepare insert statement
            insert_sql = """
                INSERT INTO pensioner_bank_master (
                    gcode, escroll_cat, gid, pension_type, branch_code, branch_name, branch_postcode,
                    branch_state, birth_year, submitted_status, waiver_upto, submission_mode,
                    verification_type, certificate_submission_date, pensioner_postcode, 
                    pensioner_distname, state, data_source, sheet_name
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """
            
            # Process data rows
            imported_count = 0
            error_count = 0
            
            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
                try:
                    # Create record with mapped values
                    record = {
                        'gcode': '',
                        'escroll_cat': '',
                        'gid': '',
                        'pension_type': '',
                        'branch_code': '',
                        'branch_name': '',
                        'branch_postcode': '',
                        'branch_state': '',
                        'birth_year': '',
                        'submitted_status': '',
                        'waiver_upto': '',
                        'submission_mode': '',
                        'verification_type': '',
                        'certificate_submission_date': '',
                        'pensioner_postcode': '',
                        'pensioner_distname': '',
                        'state': ''
                    }
                    
                    # Map values from Excel row
                    for col_idx, field_name in header_map.items():
                        if col_idx < len(row) and row[col_idx] is not None:
                            record[field_name] = str(row[col_idx]).strip()
                    
                    # Insert into database
                    values = (
                        record['gcode'],
                        record['escroll_cat'],
                        record['gid'],
                        record['pension_type'],
                        record['branch_code'],
                        record['branch_name'],
                        record['branch_postcode'],
                        record['branch_state'],
                        record['birth_year'],
                        record['submitted_status'],
                        record['waiver_upto'],
                        record['submission_mode'],
                        record['verification_type'],
                        record['certificate_submission_date'],
                        record['pensioner_postcode'],
                        record['pensioner_distname'],
                        record['state'],
                        source_name,
                        'Export Worksheet'
                    )
                    
                    self.cursor.execute(insert_sql, values)
                    imported_count += 1
                    
                    # Show progress
                    if imported_count % 5000 == 0:
                        print(f"   ✅ Imported {imported_count} records...")
                        self.conn.commit()
                        
                except Exception as e:
                    error_count += 1
                    if error_count < 10:  # Only show first 10 errors
                        print(f"   ❌ Error importing row {row_num}: {e}")
            
            self.conn.commit()
            print(f"✅ Completed {source_name} import: {imported_count} records imported, {error_count} errors")
            return imported_count
            
        except Exception as e:
            print(f"❌ Error importing {source_name}: {e}")
            return 0
            
    def import_ubi_data(self, file_path, source_name, sheet_name='Sheet2'):
        """Import UBI data"""
        print(f"\n🔄 Importing {source_name} from {file_path}")
        
        if not os.path.exists(file_path):
            print(f"❌ File not found: {file_path}")
            return 0
            
        try:
            wb = load_workbook(file_path, read_only=True)
            ws_name = sheet_name if sheet_name in wb.sheetnames else wb.sheetnames[0]
            ws = wb[ws_name]
            
            # Mapping for UBI data
            headers = [str(cell) if cell is not None else "" for cell in next(ws.iter_rows(values_only=True))]
            print(f"📋 Headers: {headers}")
            
            # Create mapping from column headers to database fields
            header_map = {}
            for idx, header in enumerate(headers):
                header_clean = header.strip().upper()
                if header_clean == 'S. NO' or header_clean == 'S NO':
                    header_map[idx] = 's_no'
                elif 'PPO' in header_clean and 'NO' in header_clean:
                    header_map[idx] = 'ppo_number'
                elif 'DATE' in header_clean and 'BIRTH' in header_clean:
                    header_map[idx] = 'pensioner_dob'
                elif header_clean == 'PSA':
                    header_map[idx] = 'psa'
                elif header_clean == 'PDA':
                    header_map[idx] = 'pda'
                elif 'BANK' in header_clean and 'DISBURSING' in header_clean:
                    header_map[idx] = 'name_of_disbursing_bank'
                elif 'BANK' in header_clean and 'BRANCH' in header_clean:
                    header_map[idx] = 'name_of_bank_branch_of_pensioner'
                elif 'PINCODE' in header_clean and 'BRANCH' not in header_clean:
                    header_map[idx] = 'pensioner_postcode'
                elif 'PENSIONER' in header_clean and 'CITY' in header_clean:
                    header_map[idx] = 'pensioner_city'
                elif header_clean == 'STATE':
                    header_map[idx] = 'state'
                elif 'PENSIONER' in header_clean and 'PINCODE' in header_clean:
                    header_map[idx] = 'pensioner_postcode'
            
            print(f"🗺️  Column mapping: {header_map}")
            
            # Prepare insert statement
            insert_sql = """
                INSERT INTO pensioner_bank_master (
                    s_no, ppo_number, pensioner_dob, psa, pda, name_of_disbursing_bank,
                    name_of_bank_branch_of_pensioner, pensioner_postcode, pensioner_city, 
                    state, data_source, sheet_name
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """
            
            # Process data rows
            imported_count = 0
            error_count = 0
            
            # Skip header rows for UBI 1 (which has extra header rows)
            start_row = 5 if 'UBI 1' in source_name else 2
            
            for row_num, row in enumerate(ws.iter_rows(min_row=start_row, values_only=True), start_row):
                try:
                    # Create record with mapped values
                    record = {
                        's_no': '',
                        'ppo_number': '',
                        'pensioner_dob': '',
                        'psa': '',
                        'pda': '',
                        'name_of_disbursing_bank': '',
                        'name_of_bank_branch_of_pensioner': '',
                        'pensioner_postcode': '',
                        'pensioner_city': '',
                        'state': ''
                    }
                    
                    # Map values from Excel row
                    for col_idx, field_name in header_map.items():
                        if col_idx < len(row) and row[col_idx] is not None:
                            record[field_name] = str(row[col_idx]).strip()
                    
                    # Insert into database
                    values = (
                        record['s_no'],
                        record['ppo_number'],
                        record['pensioner_dob'],
                        record['psa'],
                        record['pda'],
                        record['name_of_disbursing_bank'],
                        record['name_of_bank_branch_of_pensioner'],
                        record['pensioner_postcode'],
                        record['pensioner_city'],
                        record['state'],
                        source_name,
                        ws_name
                    )
                    
                    self.cursor.execute(insert_sql, values)
                    imported_count += 1
                    
                    # Show progress
                    if imported_count % 5000 == 0:
                        print(f"   ✅ Imported {imported_count} records...")
                        self.conn.commit()
                        
                except Exception as e:
                    error_count += 1
                    if error_count < 10:  # Only show first 10 errors
                        print(f"   ❌ Error importing row {row_num}: {e}")
            
            self.conn.commit()
            print(f"✅ Completed {source_name} import: {imported_count} records imported, {error_count} errors")
            return imported_count
            
        except Exception as e:
            print(f"❌ Error importing {source_name}: {e}")
            return 0

def main():
    print("🚀 STARTING ADDITIONAL DATA IMPORT")
    print("="*80)
    
    # Initialize importer
    db_path = "/data1/jainendra/DLC_backend-main/DLC_Database.db"
    importer = AdditionalDataImporter(db_path)
    
    try:
        importer.connect_db()
        
        base_path = "/data1/jainendra/DLC_backend-main"
        total_imported = 0
        
        # Import BOB data files
        bob_files = [
            ("BOB Pensioners data 1.xlsx", "BOB_DATA_1"),
            ("BOB Pensioners data 2.xlsx", "BOB_DATA_2")
        ]
        
        for file_name, source_name in bob_files:
            file_path = os.path.join(base_path, file_name)
            count = importer.import_bob_data(file_path, source_name)
            total_imported += count
            
        # Import Dashboard DLC data
        dlc_file = "Dashborad_DLC_Data_.xlsx"
        dlc_path = os.path.join(base_path, dlc_file)
        count = importer.import_dlc_data(dlc_path, "DASHBOARD_DLC_DATA")
        total_imported += count
        
        # Import UBI data files
        ubi_files = [
            ("Data from UBI 1.xlsx", "UBI_DATA_1", "Sheet1"),
            ("Data from UBI 2.xlsx", "UBI_DATA_2", "Sheet2"),
            ("Data from UBI 3.xlsx", "UBI_DATA_3", "Sheet1")
        ]
        
        for file_name, source_name, sheet_name in ubi_files:
            file_path = os.path.join(base_path, file_name)
            count = importer.import_ubi_data(file_path, source_name, sheet_name)
            total_imported += count
            
        print(f"\n🎉 IMPORT COMPLETED SUCCESSFULLY!")
        print(f"📊 Total records imported: {total_imported:,}")
        
    except Exception as e:
        print(f"❌ Import process failed: {e}")
    finally:
        importer.close_db()

if __name__ == "__main__":
    main()