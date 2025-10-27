

import sqlite3
import pandas as pd
import json
import argparse
import re
from pathlib import Path
from openpyxl import load_workbook
from tqdm import tqdm


def load_config(config_path):
    """Load the configuration JSON file."""
    with open(config_path, 'r') as f:
        return json.load(f)


def is_csv_file(file_path):
    """Check if the file is a CSV file based on extension."""
    return Path(file_path).suffix.lower() == '.csv'


def read_excel(excel_path, sheet_name=0):
    """Read entire Excel sheet into a DataFrame."""
    return pd.read_excel(excel_path, sheet_name=sheet_name)


def read_csv_streaming(csv_path, chunk_size=1000):
    """Stream rows from CSV file in chunks."""
    for chunk in pd.read_csv(csv_path, chunksize=chunk_size):
        yield chunk


def read_excel_streaming(excel_path, sheet_name=0, chunk_size=1000):
    """Stream rows from Excel sheet in chunks using openpyxl."""
    wb = load_workbook(filename=excel_path, read_only=True, data_only=True)

    if isinstance(sheet_name, int):
        ws = wb.worksheets[sheet_name]
    else:
        ws = wb[sheet_name]

    rows_iter = ws.iter_rows(values_only=True)
    header = next(rows_iter)
    header = [col if col is not None else f"Column_{i}" for i, col in enumerate(header)]

    chunk = []
    for row in rows_iter:
        # Skip completely empty rows
        if all(cell is None or str(cell).strip() == '' for cell in row):
            continue
        
        chunk.append(row)
        if len(chunk) >= chunk_size:
            yield pd.DataFrame(chunk, columns=header)
            chunk = []

    if chunk:
        yield pd.DataFrame(chunk, columns=header)

    wb.close()


def clean_pincode(pincode_value):
    """Extract numeric pincode from strings that may contain address information.
    
    Examples:
        'Bongaigaon S.O , Pin- 783380' -> '783380'
        '783380' -> '783380'
        'Pin: 123456' -> '123456'
        'Address, Pincode-654321' -> '654321'
    """
    if pd.isna(pincode_value):
        return None
    
    pincode_str = str(pincode_value).strip()
    
    # If it's already a clean 6-digit pincode, return it
    if pincode_str.isdigit() and len(pincode_str) == 6:
        return pincode_str
    
    # Extract 6-digit pincode using regex
    # Look for patterns like "Pin- 123456", "Pin: 123456", "Pincode-123456", or just "123456"
    pincode_match = re.search(r'\b(\d{6})\b', pincode_str)
    if pincode_match:
        return pincode_match.group(1)
    
    # If no 6-digit pincode found, try to extract any digits and pad/truncate to 6
    digits_only = re.sub(r'\D', '', pincode_str)
    if len(digits_only) >= 6:
        return digits_only[:6]
    elif len(digits_only) > 0:
        # Return as is if less than 6 digits (might be invalid but preserve it)
        return digits_only
    
    return None


def handle_null_values(df, null_handling_config=None):
    """Handle NULL values based on config."""
    if not null_handling_config:
        return df
    for col, val in null_handling_config.items():
        if col in df.columns:
            df[col] = df[col].fillna(val)
    return df


def create_table_if_not_exists(conn, table_name, df):
    """Create SQLite table automatically if it doesn't exist."""
    cursor = conn.cursor()
    columns_sql = []
    for col, dtype in df.dtypes.items():
        if pd.api.types.is_integer_dtype(dtype):
            col_type = "INTEGER"
        elif pd.api.types.is_float_dtype(dtype):
            col_type = "REAL"
        else:
            col_type = "TEXT"
        columns_sql.append(f'"{col}" {col_type}')
    sql = f'CREATE TABLE IF NOT EXISTS "{table_name}" ({", ".join(columns_sql)})'
    cursor.execute(sql)
    conn.commit()
    cursor.close()


def insert_data_streaming_all_sheets(db_path, table_name, excel_path, column_mapping,
                                    static_fields=None, chunk_size=1000,
                                    null_handling_config=None):
    """Insert all sheets from workbook or CSV file into DB with dynamic sheet_name and file_name."""
    # Extract filename from path
    file_name = Path(excel_path).name
    
    # Check if it's a CSV file
    if is_csv_file(excel_path):
        print(f"📄 Detected CSV file: {file_name}")
        sheet_names = ["Sheet1"]  # CSV files have only one "sheet"
    else:
        # It's an Excel file
        wb = load_workbook(filename=excel_path, read_only=True, data_only=True)
        sheet_names = wb.sheetnames
        wb.close()
        print(f"📚 Found {len(sheet_names)} sheets: {sheet_names}")

    print(f"📁 File name: {file_name}")
    total_inserted_all = 0

    for sheet_name in sheet_names:
        print(f"\n📄 Processing sheet: {sheet_name}")
        total_inserted = 0
        conn = sqlite3.connect(db_path)
        cursor = conn.cursor()

        def _build_insert_query(table, columns):
            cols_sql = ', '.join([f'"{c}"' for c in columns])
            placeholders = ', '.join(['?'] * len(columns))
            return f'INSERT INTO "{table}" ({cols_sql}) VALUES ({placeholders})'

        try:
            reverse_mapping = {v: k for k, v in column_mapping.items()}
            excel_columns = list(reverse_mapping.keys())

            chunk_count = 0
            
            # Choose the appropriate streaming function based on file type
            if is_csv_file(excel_path):
                chunk_iterator = read_csv_streaming(excel_path, chunk_size)
            else:
                chunk_iterator = read_excel_streaming(excel_path, sheet_name, chunk_size)
            
            for chunk_df in tqdm(chunk_iterator,
                                desc=f"Reading '{sheet_name}'", unit="chunk"):
                available_columns = [col for col in excel_columns if col in chunk_df.columns]
                df_mapped = chunk_df[available_columns].copy()
                df_mapped.rename(columns=reverse_mapping, inplace=True)

                # Add static fields
                if static_fields:
                    for db_col, val in static_fields.items():
                        df_mapped[db_col] = val
                
                # Handle file_name and sheet_name
                # If file_name column exists in the data (mapped), use it; otherwise use the actual filename
                if "file_name" not in df_mapped.columns:
                    df_mapped["file_name"] = file_name
                
                # If sheet_name column exists in the data (mapped), use it; otherwise use the actual sheet name
                if "sheet_name" not in df_mapped.columns:
                    df_mapped["sheet_name"] = sheet_name

                # Clean pincode fields to extract only numeric values
                pincode_columns = ['Pensioner_pincode', 'Branch_pincode']
                for pincode_col in pincode_columns:
                    if pincode_col in df_mapped.columns:
                        if chunk_count == 0:
                            sample_before = df_mapped[pincode_col].head(3).tolist()
                        
                        df_mapped[pincode_col] = df_mapped[pincode_col].apply(clean_pincode)
                        
                        if chunk_count == 0:
                            sample_after = df_mapped[pincode_col].head(3).tolist()
                            print(f"📍 {pincode_col} cleaning - Sample before: {sample_before}, after: {sample_after}")
                
                # Use Branch_pincode as fallback for NULL Pensioner_pincode
                if 'Pensioner_pincode' in df_mapped.columns and 'Branch_pincode' in df_mapped.columns:
                    mask = df_mapped['Pensioner_pincode'].isna()
                    if mask.any():
                        fallback_count = mask.sum()
                        if chunk_count == 0:
                            print(f"📍 Using Branch_pincode as fallback for {fallback_count} NULL Pensioner_pincode values in first chunk")
                        df_mapped.loc[mask, 'Pensioner_pincode'] = df_mapped.loc[mask, 'Branch_pincode']

                # Extract year from YOB/DOB field if it contains full date
                if 'YOB' in df_mapped.columns:
                    def extract_year(dob_value):
                        """Extract year from DOB string in various formats."""
                        if pd.isna(dob_value):
                            return None
                        
                        dob_str = str(dob_value).strip()
                        
                        # If it's already a 4-digit year, return it
                        if dob_str.isdigit() and len(dob_str) == 4:
                            return int(dob_str)
                        
                        # Try to parse as datetime and extract year
                        try:
                            # Try common date formats
                            for fmt in ['%d-%m-%Y', '%d/%m/%Y', '%Y-%m-%d', '%Y/%m/%d', '%m-%d-%Y', '%m/%d/%Y']:
                                try:
                                    dt = pd.to_datetime(dob_str, format=fmt)
                                    return dt.year
                                except:
                                    continue
                            
                            # If no specific format worked, let pandas infer
                            dt = pd.to_datetime(dob_str, errors='coerce')
                            if pd.notna(dt):
                                return dt.year
                        except:
                            pass
                        
                        # Try to extract 4-digit year using regex
                        import re
                        year_match = re.search(r'\b(19\d{2}|20\d{2})\b', dob_str)
                        if year_match:
                            return int(year_match.group(1))
                        
                        return None
                    
                    if chunk_count == 0:
                        sample_before = df_mapped['YOB'].head(3).tolist()
                    
                    df_mapped['YOB'] = df_mapped['YOB'].apply(extract_year)
                    
                    if chunk_count == 0:
                        sample_after = df_mapped['YOB'].head(3).tolist()
                        print(f"📅 YOB extraction - Sample before: {sample_before}, after: {sample_after}")

                # Handle NULLs
                df_mapped = handle_null_values(df_mapped, null_handling_config)
                
                # Skip empty chunks (after all transformations)
                if df_mapped.empty or len(df_mapped) == 0:
                    continue

                # Auto-create table if it doesn't exist
                if chunk_count == 0 and chunk_df.shape[0] > 0:
                    create_table_if_not_exists(conn, table_name, df_mapped)

                insert_query = _build_insert_query(table_name, list(df_mapped.columns))
                rows_in_chunk = len(df_mapped)

                # Use executemany for batch insertion (much faster than row-by-row)
                rows_data = df_mapped.values.tolist()
                cursor.executemany(insert_query, rows_data)
                total_inserted += rows_in_chunk

                conn.commit()
                chunk_count += 1
                print(f"  ✓ Inserted {rows_in_chunk} rows from chunk {chunk_count} ({sheet_name})")

            total_inserted_all += total_inserted
            print(f"✅ {total_inserted} rows inserted from sheet '{sheet_name}'")

        except Exception as e:
            print(f"\n❌ Error in sheet '{sheet_name}': {e}")
            conn.rollback()
        finally:
            cursor.close()
            conn.close()

    print(f"\n🎯 All sheets processed. Total rows inserted: {total_inserted_all}")


def main():
    parser = argparse.ArgumentParser(description="Import CSV or Excel files into SQLite database")
    parser.add_argument('excel_file', help='Path to CSV file (.csv) or Excel workbook (.xlsx)')
    parser.add_argument('config_file', help='Path to config JSON file (contains column_mapping)')
    parser.add_argument('db_file', help='Path to SQLite database')
    parser.add_argument('table_name', help='Target table name')
    parser.add_argument('--chunk-size', type=int, default=1000, help='Rows per chunk (default: 1000)')
    parser.add_argument('--stream', action='store_true', help='Enable streaming mode for large files')
    parser.add_argument('--static-field', '-s', action='append', nargs=2, metavar=('COLUMN', 'VALUE'),
                        help='Add static field (e.g., -s file_name "Batch1")')
    parser.add_argument('--fill-null', '-f', action='append', nargs=2, metavar=('COLUMN', 'VALUE'),
                        help='Fill NULL values (e.g., -f Pensioner_pincode 0)')
    args = parser.parse_args()

    if not Path(args.excel_file).exists():
        print(f"❌ File not found: {args.excel_file}")
        return
    if not Path(args.config_file).exists():
        print(f"❌ Config file not found: {args.config_file}")
        return

    print(f"📖 Loading config from {args.config_file}...")
    config = load_config(args.config_file)
    if 'column_mapping' not in config:
        print("❌ Config must contain 'column_mapping' key.")
        return
    column_mapping = config['column_mapping']

    static_fields = {}
    if args.static_field:
        for col, val in args.static_field:
            static_fields[col] = val

    null_handling_config = {}
    if args.fill_null:
        for col, val in args.fill_null:
            try:
                null_handling_config[col] = int(val)
            except ValueError:
                try:
                    null_handling_config[col] = float(val)
                except ValueError:
                    null_handling_config[col] = val

    print(f"🧩 Static fields: {static_fields}")
    print(f"🔧 NULL handling: {null_handling_config}")
    
    # Determine file type
    file_type = "CSV" if is_csv_file(args.excel_file) else "Excel"
    print(f"📦 Starting import of {file_type} file into '{args.table_name}'...")

    insert_data_streaming_all_sheets(
        args.db_file,
        args.table_name,
        args.excel_file,
        column_mapping,
        static_fields,
        args.chunk_size,
        null_handling_config
    )

    print(f"\n✅ Done! {file_type} file imported successfully.")


if __name__ == "__main__":
    main()
